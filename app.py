import hashlib
import re
from datetime import date, datetime
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from supabase_backend import (
    all_daily, create_station, daily_rows, delete_daily, get_supabase, healthcheck,
    hourly_rows, month_daily, save_daily, stations, update_station,
)

APP_TITLE = "蒙西新能源多场站日清分与交易决策系统"
st.set_page_config(page_title=APP_TITLE, page_icon="⚡", layout="wide")

RULE_VERSION = "蒙西2026（自动结算V2.4）"
BASELINE_PRICE = 282.9
MONTH_LOWER = 0.90
MONTH_UPPER = 1.10
LOWER_FACTOR = 1.30
UPPER_FACTOR = 1.10
RISK_COMP_BASE = 0.50
RISK_REC_BASE = 1.45
CURVE_LINK = 0.50


def fnum(v):
    try:
        return float(str(v).replace(",", "")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0


def wavg(v, w):
    v, w = np.asarray(v, float), np.asarray(w, float)
    ok = np.isfinite(v) & np.isfinite(w)
    if not ok.any():
        return 0.0
    sw = np.sum(w[ok])
    return float(np.sum(v[ok] * w[ok]) / sw) if sw else float(np.mean(v[ok]))


def money(v):
    return f"{float(v or 0):,.2f}"


def detect_date(name, ws):
    pats = [r"(20\d{2})[-_/年](\d{1,2})[-_/月](\d{1,2})", r"(20\d{2})(\d{2})(\d{2})"]
    texts = [name]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        for v in row:
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, str):
                texts.append(v)
    for text in texts:
        for p in pats:
            m = re.search(p, text)
            if m:
                try:
                    return date(*map(int, m.groups()))
                except ValueError:
                    pass
    return None


def detect_station(filename):
    df = stations(False)
    if df.empty:
        return None, None
    fn, hits = filename.lower(), []
    for _, r in df.iterrows():
        for text in (r.get("name"), r.get("short_name")):
            if isinstance(text, str) and text.strip() and text.strip().lower() in fn:
                hits.append((len(text.strip()), int(r["id"]), r["name"]))
    if not hits:
        return None, None
    _, sid, name = max(hits)
    return sid, name


def parse_excel(data, filename):
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    trade_date = detect_date(filename, ws)
    best = None
    max_start = max(2, min(ws.max_row, 500) - 95)
    for start in range(1, max_start):
        score = 0
        for r in range(start, start + 96):
            score += sum(isinstance(ws.cell(r, c).value, (int, float)) for c in (3, 4, 10, 11, 12))
        if score >= 288 and (best is None or score > best[0]):
            best = (score, start)
    if not best:
        raise ValueError("未识别到96点数据；要求 C/D/J/K/L 为中长期电力/中长期价/实时价/实际计量/统一结算点价。")

    start = best[1]
    raw = pd.DataFrame([
        {"lt_power": fnum(ws.cell(r, 3).value), "lt_price": fnum(ws.cell(r, 4).value),
         "spot_price": fnum(ws.cell(r, 10).value), "actual_power": fnum(ws.cell(r, 11).value),
         "unified_price": fnum(ws.cell(r, 12).value)}
        for r in range(start, start + 96)
    ])
    rows = []
    for h in range(24):
        g = raw.iloc[h * 4:(h + 1) * 4]
        le, lp, ae = float(g.lt_power.mean()), float(g.lt_price.mean()), float(g.actual_power.mean())
        sp, up = wavg(g.spot_price, g.actual_power), float(g.unified_price.mean())
        sf, lf = ae * sp, le * (lp - up)
        rows.append([h + 1, le, lp, ae, sp, up, sf, lf, sf + lf])
    hourly = pd.DataFrame(rows, columns=["时点", "中长期电量", "中长期价", "上网电量", "实时价", "统一结算点价", "现货电费", "中长期差价电费", "电能量合计"])
    summary = {
        "trade_date": trade_date,
        "lt_energy": hourly["中长期电量"].sum(),
        "lt_price": wavg(hourly["中长期价"], hourly["中长期电量"]),
        "actual_energy": hourly["上网电量"].sum(),
        "spot_price": wavg(hourly["实时价"], hourly["上网电量"]),
        "unified_price": wavg(hourly["统一结算点价"], hourly["中长期电量"]),
        "spot_fee": hourly["现货电费"].sum(),
        "lt_diff_fee": hourly["中长期差价电费"].sum(),
        "energy_total": hourly["电能量合计"].sum(),
    }
    return summary, hourly, start


def month_summary(df):
    if df.empty:
        return {}
    x = df.copy()
    for c in ["lt_energy", "lt_price", "actual_energy", "spot_price", "unified_price", "energy_total"]:
        if c not in x.columns:
            x[c] = 0.0
        x[c] = pd.to_numeric(x[c], errors="coerce").fillna(0.0)
    lt, act = float(x.lt_energy.sum()), float(x.actual_energy.sum())
    return {
        "lt": lt, "actual": act, "coverage": lt / act if act else 0.0,
        "lt_price": wavg(x.lt_price, x.lt_energy),
        "spot_price": wavg(x.spot_price, x.actual_energy),
        "unified_price": wavg(x.unified_price, x.lt_energy),
        "energy": float(x.energy_total.sum()),
    }


def station_selector(key, active=True):
    sdf = stations(active)
    if sdf.empty:
        st.warning("请先在“场站管理”创建场站。")
        return None, sdf
    name = st.selectbox("场站", sdf["name"].tolist(), key=key)
    return sdf[sdf.name == name].iloc[0], sdf


def load_month_input(station_id, month):
    data = get_supabase().table("monthly_settlement").select("*").eq("station_id", station_id).eq("trade_month", month).limit(1).execute().data or []
    return data[0] if data else {}


def save_month_input(station_id, month, payload):
    row = {"station_id": station_id, "trade_month": month, "rule_version": RULE_VERSION, **payload}
    get_supabase().table("monthly_settlement").upsert(row, on_conflict="station_id,trade_month").execute()


def auto_settlement(sm, inp):
    if not sm:
        return {}
    lt, act, p_lt, p_spot, energy = sm["lt"], sm["actual"], sm["lt_price"], sm["spot_price"], sm["energy"]
    p_unified = p_spot
    p_node = p_spot
    p_regional = p_lt

    me = max(0.0, float(inp.get("mechanism_energy") or 0.0))
    assess_actual = max(0.0, act - me)
    assess_coverage = lt / assess_actual if assess_actual > 0 else 0.0
    lower_energy, upper_energy = assess_actual * MONTH_LOWER, assess_actual * MONTH_UPPER

    upper_assess = 0.0
    if assess_actual > 0 and assess_coverage > MONTH_UPPER and p_lt > p_unified:
        upper_assess = max(0.0, (lt - upper_energy) * (p_lt * UPPER_FACTOR - p_unified))

    lower_assess = 0.0
    lower_ready = p_regional > 0
    if lower_ready and assess_actual > 0 and assess_coverage < MONTH_LOWER and p_node > p_regional:
        lower_assess = max(0.0, (lower_energy - lt) * (p_node * LOWER_FACTOR - p_regional))
    assessment = max(upper_assess, lower_assess)

    market_avg = float(inp.get("market_bilateral_listing_avg") or 0.0)
    curve = float(inp.get("curve_reasonability") or 0.0)
    risk_ready = market_avg > 0 and curve > 0
    comp_ratio = max(0.0, RISK_COMP_BASE - (1 - curve) * CURVE_LINK) if risk_ready else 0.0
    rec_ratio = RISK_REC_BASE + (1 - curve) * CURVE_LINK if risk_ready else 0.0
    congestion = float(inp.get("congestion") or 0.0)
    pre_other = float(inp.get("pre_risk_other_fee") or 0.0)
    pre_risk = energy + congestion + pre_other - assessment
    risk = 0.0
    if risk_ready and market_avg <= BASELINE_PRICE and act > 0 and p_lt > 0:
        floor, cap = act * p_lt * comp_ratio, act * p_lt * rec_ratio
        if pre_risk < floor:
            risk = floor - pre_risk
        elif pre_risk > cap:
            risk = cap - pre_risk

    # 绿电口径：取“绿电合约电量”和“上网电量-机制电量”的较小值。
    ge = max(0.0, float(inp.get("green_contract_energy") or 0.0))
    gp = float(inp.get("green_environment_price") or 0.0)
    green_available = max(0.0, act - me)
    green_energy = min(ge, green_available)
    green_fee = green_energy * gp

    mp = float(inp.get("mechanism_price") or BASELINE_PRICE)
    ms = float(inp.get("mechanism_spot_price") or 0.0)
    mechanism_fee = me * (mp - ms) if me > 0 and ms != 0 else 0.0
    regular = float(inp.get("regular_fee") or 0.0)
    unit_fee = float(inp.get("unit_fee") or 0.0)
    manual = float(inp.get("manual_adjustment") or 0.0)
    final = energy + congestion - assessment + risk + green_fee + mechanism_fee + regular - unit_fee + manual
    return {
        "assessment_actual": assess_actual, "assessment_coverage": assess_coverage,
        "assessment_lower_energy": lower_energy, "assessment_upper_energy": upper_energy,
        "p_lt": p_lt, "p_unified": p_unified, "p_node": p_node, "p_regional": p_regional,
        "upper_assessment": upper_assess, "lower_assessment": lower_assess, "lower_ready": lower_ready,
        "assessment": assessment, "risk_ready": risk_ready, "risk_prevention": risk,
        "pre_risk_revenue": pre_risk, "green_contract_energy": ge, "green_available_energy": green_available,
        "green_energy": green_energy, "green_environment_price": gp, "green_fee": green_fee,
        "mechanism_fee": mechanism_fee, "final_revenue": final, "final_price": final / act if act else 0.0,
    }


try:
    healthcheck()
    DB_OK, DB_ERROR = True, ""
except Exception as exc:
    DB_OK, DB_ERROR = False, str(exc)

st.sidebar.title("⚡ 蒙西交易系统")
st.sidebar.caption("Supabase 云数据库版 · 自动结算V2.4")
if DB_OK:
    st.sidebar.success("数据库已连接")
else:
    st.sidebar.error("数据库未连接")
page = st.sidebar.radio("导航", ["总览", "场站管理", "日清分上传", "月度累计", "自动结算", "交易决策", "数据管理", "系统状态"])
if not DB_OK and page != "系统状态":
    st.error("Supabase 未连接，请进入系统状态查看错误。")
    st.stop()

if page == "总览":
    st.title(APP_TITLE)
    sdf, daily = stations(False), all_daily()
    c = st.columns(4)
    c[0].metric("场站数", len(sdf)); c[1].metric("累计日清分", len(daily))
    c[2].metric("累计上网电量", f"{pd.to_numeric(daily.get('actual_energy', pd.Series(dtype=float)), errors='coerce').fillna(0).sum():,.2f} MWh")
    c[3].metric("结算引擎", RULE_VERSION)
    st.info("创建场站 → 批量上传日清分 → 一键全部入库 → 月度累计 → 自动结算 → 交易决策。")
    if not daily.empty:
        st.dataframe(daily.head(30), use_container_width=True, hide_index=True)

elif page == "场站管理":
    st.title("场站管理")
    with st.form("new_station"):
        a, b, c = st.columns(3)
        name = a.text_input("场站名称 *")
        short = b.text_input("识别关键词", help="例如：战壕梁")
        typ = c.selectbox("类型", ["光伏", "风电", "其他"])
        a, b, c, d = st.columns(4)
        cap = a.number_input("装机 MW", 0.0, 10000.0, 50.0)
        upper = b.number_input("仓位风险上限 %", 0.0, 300.0, 110.0) / 100
        lower = c.number_input("目标仓位下限 %", 0.0, 300.0, 80.0) / 100
        limit = d.number_input("单次交易上限 MWh", 0.0, 100000.0, 100.0)
        spread = st.number_input("最小价差阈值 元/MWh", 0.0, 1000.0, 10.0)
        if st.form_submit_button("创建场站", type="primary"):
            if not name.strip():
                st.error("场站名称不能为空")
            else:
                create_station({"name": name.strip(), "short_name": short.strip() or name.strip(), "station_type": typ, "capacity_mw": cap, "risk_upper": upper, "target_lower": lower, "trade_limit_mwh": limit, "min_spread": spread, "active": True})
                st.success("已创建"); st.rerun()
    sdf = stations(False)
    if not sdf.empty:
        st.dataframe(sdf, use_container_width=True, hide_index=True)
        with st.expander("编辑场站"):
            en = st.selectbox("场站", sdf.name.tolist())
            r = sdf[sdf.name == en].iloc[0]
            kw = st.text_input("识别关键词", str(r.get("short_name") or r["name"]))
            active = st.checkbox("启用", bool(r.get("active", True)))
            if st.button("保存"):
                update_station(int(r.id), {"short_name": kw.strip(), "active": active})
                st.success("已保存"); st.rerun()

elif page == "日清分上传":
    st.title("日清分批量上传")
    st.caption("可一次选择一个月多个场站的日清分。系统先自动识别并预览，再一键全部入库。")
    sdf = stations()
    if sdf.empty:
        st.warning("请先创建场站"); st.stop()
    names, ids = sdf.name.tolist(), dict(zip(sdf.name, sdf.id))
    files = st.file_uploader("选择多个 Excel 文件", type=["xlsx", "xlsm"], accept_multiple_files=True)
    overwrite = st.checkbox("覆盖数据库中同场站、同日期已有数据", value=False)
    parsed, preview_rows = [], []
    for i, f in enumerate(files or []):
        try:
            data = f.getvalue(); summary, hourly, start = parse_excel(data, f.name); _, auto_name = detect_station(f.name)
            station_name, trade_date = auto_name or "", summary.get("trade_date")
            status = "可入库" if station_name and trade_date else ("未识别场站" if not station_name else "未识别日期")
            parsed.append({"index": i, "file": f, "data": data, "summary": summary, "hourly": hourly, "station_name": station_name, "trade_date": trade_date, "start": start, "error": None})
            preview_rows.append({"文件名": f.name, "自动识别场站": station_name or "未识别", "日期": str(trade_date or "未识别"), "中长期电量": round(float(summary["lt_energy"]), 3), "上网电量": round(float(summary["actual_energy"]), 3), "电能量合计": round(float(summary["energy_total"]), 2), "状态": status})
        except Exception as exc:
            parsed.append({"index": i, "file": f, "error": str(exc)})
            preview_rows.append({"文件名": f.name, "自动识别场站": "-", "日期": "-", "中长期电量": 0, "上网电量": 0, "电能量合计": 0, "状态": f"解析失败：{exc}"})
    if preview_rows:
        preview = pd.DataFrame(preview_rows)
        st.subheader(f"识别预览｜共 {len(preview)} 个文件"); st.dataframe(preview, use_container_width=True, hide_index=True)
        ready_count = int((preview["状态"] == "可入库").sum())
        a, b, c = st.columns(3); a.metric("可直接入库", ready_count); b.metric("需处理", len(preview) - ready_count); c.metric("总文件数", len(preview))
        if st.button("🚀 一键全部入库", type="primary", use_container_width=True):
            success, skipped, failed = [], [], []; progress = st.progress(0, text="开始入库..."); total = max(len(parsed), 1)
            for pos, item in enumerate(parsed, start=1):
                f = item["file"]
                if item.get("error"):
                    failed.append((f.name, item["error"]))
                elif not item.get("station_name") or not item.get("trade_date"):
                    skipped.append((f.name, "未识别场站或日期"))
                else:
                    try:
                        summary = item["summary"]; summary["trade_date"] = item["trade_date"]
                        save_daily(int(ids[item["station_name"]]), f.name, hashlib.md5(item["data"]).hexdigest(), summary, item["hourly"], overwrite)
                        success.append(f.name)
                    except ValueError as exc:
                        skipped.append((f.name, str(exc)))
                    except Exception as exc:
                        failed.append((f.name, str(exc)))
                progress.progress(pos / total, text=f"正在处理 {pos}/{total}")
            progress.empty(); st.success(f"批量入库完成：成功 {len(success)} 个，跳过 {len(skipped)} 个，失败 {len(failed)} 个。")
            if skipped: st.dataframe(pd.DataFrame(skipped, columns=["跳过文件", "原因"]), use_container_width=True, hide_index=True)
            if failed: st.dataframe(pd.DataFrame(failed, columns=["失败文件", "错误"]), use_container_width=True, hide_index=True)
        with st.expander("单文件检查/手动入库"):
            for i, item in enumerate(parsed):
                if item.get("error"): continue
                f = item["file"]; st.markdown(f"**{f.name}**"); a, b = st.columns(2)
                auto_name = item.get("station_name"); idx = names.index(auto_name) if auto_name in names else 0
                station_name = a.selectbox("归属场站", names, index=idx, key=f"manual_station_{i}")
                trade_date = b.date_input("交易日期", item.get("trade_date") or date.today(), key=f"manual_date_{i}")
                if st.button("单独入库", key=f"manual_save_{i}"):
                    summary = item["summary"]; summary["trade_date"] = trade_date
                    try:
                        save_daily(int(ids[station_name]), f.name, hashlib.md5(item["data"]).hexdigest(), summary, item["hourly"], overwrite); st.success(f"{f.name} 已入库")
                    except Exception as exc:
                        st.error(str(exc))
                with st.expander(f"查看24时点：{f.name}"): st.dataframe(item["hourly"], use_container_width=True, hide_index=True)

elif page == "月度累计":
    st.title("月度累计")
    row, _ = station_selector("month_station")
    if row is None: st.stop()
    month = st.text_input("月份 YYYY-MM", date.today().strftime("%Y-%m"), key="month")
    df = month_daily(int(row.id), month); sm = month_summary(df)
    if not sm: st.info("暂无数据"); st.stop()
    c = st.columns(7)
    c[0].metric("累计上网", f"{sm['actual']:,.2f}"); c[1].metric("累计中长期", f"{sm['lt']:,.2f}"); c[2].metric("原始覆盖率", f"{sm['coverage']:.2%}")
    c[3].metric("中长期均价", f"{sm['lt_price']:.2f}"); c[4].metric("现货均价", f"{sm['spot_price']:.2f}"); c[5].metric("原表统一结算点均价", f"{sm['unified_price']:.2f}"); c[6].metric("电能量收入", money(sm['energy']))
    st.dataframe(df, use_container_width=True, hide_index=True)

elif page == "自动结算":
    st.title("自动结算｜月度结算表")
    st.caption("基准电量、价格参数、考核、风险防范、机制、绿电和最终收益集中在一张结算表中。")
    row, _ = station_selector("settle_station")
    if row is None: st.stop()
    month = st.text_input("月份 YYYY-MM", date.today().strftime("%Y-%m"), key="settle_month")
    sm = month_summary(month_daily(int(row.id), month))
    if not sm: st.info("该月暂无日清分"); st.stop()
    old = load_month_input(int(row.id), month)

    with st.expander("月度公共参数", expanded=not bool(old)):
        with st.form("market_inputs"):
            a, b = st.columns(2)
            marketavg = a.number_input("区内协商/挂牌成交加权均价", value=float(old.get("market_bilateral_listing_avg") or 0.0))
            curve = b.number_input("正式曲线合理度 %", 0.0, 100.0, value=float(old.get("curve_reasonability") or 0.0) * 100) / 100
            a, b, c = st.columns(3)
            congestion = a.number_input("阻塞盈余返还/分摊（元）", value=float(old.get("congestion") or 0.0))
            pre_other = b.number_input("风险防范前其他费用净额（元）", value=float(old.get("pre_risk_other_fee") or 0.0))
            regular = c.number_input("风险防范后其他常规净额（元）", value=float(old.get("regular_fee") or 0.0))
            st.markdown("**绿电**")
            a, b = st.columns(2)
            ge = a.number_input("绿电合约电量 MWh", value=float(old.get("green_contract_energy") or 0.0), help="绿电电量取该值与（上网电量-机制电量）的较小值。")
            gp = b.number_input("绿色权益价格 元/MWh", value=float(old.get("green_environment_price") or 0.0))
            st.markdown("**机制电量**")
            a, b, c = st.columns(3)
            me = a.number_input("机制电量 MWh", value=float(old.get("mechanism_energy") or 0.0), help="先从Qactual中扣除，再计算90%/110%边界，同时用于绿电可结算电量。")
            mp = b.number_input("机制电价 元/MWh", value=float(old.get("mechanism_price") or BASELINE_PRICE))
            ms = c.number_input("机制电量对应现货均价", value=float(old.get("mechanism_spot_price") or 0.0))
            a, b = st.columns(2)
            unit = a.number_input("机组/两个细则等扣费（元）", value=float(old.get("unit_fee") or 0.0))
            manual = b.number_input("人工最终调整（元）", value=float(old.get("manual_adjustment") or 0.0))
            if st.form_submit_button("保存月度参数并重算", type="primary"):
                save_month_input(int(row.id), month, {
                    "market_bilateral_listing_avg": marketavg, "curve_reasonability": curve,
                    "congestion": congestion, "pre_risk_other_fee": pre_other, "regular_fee": regular,
                    "green_contract_energy": ge, "green_environment_price": gp,
                    "mechanism_energy": me, "mechanism_price": mp, "mechanism_spot_price": ms,
                    "unit_fee": unit, "manual_adjustment": manual,
                })
                st.success("已保存并重算"); st.rerun()

    inp = load_month_input(int(row.id), month)
    calc = auto_settlement(sm, inp)
    me = max(0.0, float(inp.get("mechanism_energy") or 0.0))

    st.subheader(f"{row['name']}｜{month} 自动结算表")
    settle_rows = [
        ["基础数据", "实际上网电量 Qactual", "日清分月累计", sm["actual"], "MWh", "自动"],
        ["基础数据", "机制电量 Q机制", "月度机制参数", me, "MWh", "已录入" if me > 0 else "未录入/0"],
        ["考核基准", "扣机制后考核电量", "Qactual - Q机制", calc.get("assessment_actual", 0), "MWh", "自动"],
        ["考核基准", "90%下限电量", "(Qactual-Q机制)×90%", calc.get("assessment_lower_energy", 0), "MWh", "自动"],
        ["考核基准", "110%上限电量", "(Qactual-Q机制)×110%", calc.get("assessment_upper_energy", 0), "MWh", "自动"],
        ["考核基准", "中长期电量 QLT", "日清分月累计", sm["lt"], "MWh", "自动"],
        ["考核基准", "考核签约率", "QLT/(Qactual-Q机制)", calc.get("assessment_coverage", 0) * 100, "%", "自动"],
        ["价格参数", "P_LT", "本站中长期均价", calc.get("p_lt", 0), "元/MWh", "自动"],
        ["价格参数", "P_统一", "取现货均价", calc.get("p_unified", 0), "元/MWh", "自动"],
        ["价格参数", "P_节点", "取现货均价", calc.get("p_node", 0), "元/MWh", "自动"],
        ["价格参数", "P_区域同类型", "取本站中长期均价", calc.get("p_regional", 0), "元/MWh", "自动"],
        ["考核费用", "全月上限考核", "[QLT-(Qactual-Q机制)×110%]×(1.1P_LT-P_统一)", calc.get("upper_assessment", 0), "元", "自动"],
        ["考核费用", "全月下限考核", "[(Qactual-Q机制)×90%-QLT]×(1.3P_节点-P_区域同类型)", calc.get("lower_assessment", 0), "元", "自动"],
        ["考核费用", "自动考核合计", "max(上限考核, 下限考核)", calc.get("assessment", 0), "元", "自动"],
        ["电能结算", "电能量收入", "日清分电能量合计月累计", sm["energy"], "元", "自动"],
        ["风险防范", "风险防范前收入", "电能量收入+阻塞盈余+其他净额-考核", calc.get("pre_risk_revenue", 0), "元", "自动"],
        ["风险防范", "风险防范金额", "按风险防范参数计算", calc.get("risk_prevention", 0), "元", "自动" if calc.get("risk_ready") else "待公共参数"],
        ["绿电结算", "绿电合约电量", "月度录入", calc.get("green_contract_energy", 0), "MWh", "参数"],
        ["绿电结算", "绿电可结算上网电量", "max(Qactual-Q机制,0)", calc.get("green_available_energy", 0), "MWh", "自动"],
        ["绿电结算", "绿电电量", "min(绿电合约电量, Qactual-Q机制)", calc.get("green_energy", 0), "MWh", "自动"],
        ["绿电结算", "绿色权益价格", "月度录入", calc.get("green_environment_price", 0), "元/MWh", "参数"],
        ["绿电结算", "绿电费用", "绿电电量×绿色权益价格", calc.get("green_fee", 0), "元", "自动"],
        ["其他结算", "机制费用", "Q机制×(P机制-P机制现货)", calc.get("mechanism_fee", 0), "元", "自动/参数"],
        ["最终结果", "预计最终收益", "电能量±各项结算费用", calc.get("final_revenue", 0), "元", "自动"],
        ["最终结果", "最终结算均价", "预计最终收益/Qactual", calc.get("final_price", 0), "元/MWh", "自动"],
    ]
    settle_df = pd.DataFrame(settle_rows, columns=["类别", "结算项目", "计算口径/公式", "结果", "单位", "状态"])
    settle_df["结果"] = pd.to_numeric(settle_df["结果"], errors="coerce").round(4)
    st.dataframe(settle_df, use_container_width=True, hide_index=True, height=850)

    st.subheader("费用汇总")
    fee_df = pd.DataFrame([
        ["电能量收入", sm["energy"]],
        ["实际计入考核", -float(calc.get("assessment", 0))],
        ["风险防范", float(calc.get("risk_prevention", 0))],
        ["绿电费用", float(calc.get("green_fee", 0))],
        ["机制费用", float(calc.get("mechanism_fee", 0))],
        ["预计最终收益", float(calc.get("final_revenue", 0))],
    ], columns=["项目", "金额（元）"])
    fee_df["金额（元）"] = fee_df["金额（元）"].round(2)
    st.dataframe(fee_df, use_container_width=True, hide_index=True)
    st.caption("绿电口径：Q绿电=min(Q绿电合约, Qactual-Q机制)；绿电费用=Q绿电×绿色权益价格。")

elif page == "交易决策":
    st.title("月内交易决策")
    row, _ = station_selector("decision_station")
    if row is None: st.stop()
    month = st.text_input("月份 YYYY-MM", date.today().strftime("%Y-%m"), key="decision_month")
    sm = month_summary(month_daily(int(row.id), month)) or {"actual": 0, "lt": 0, "coverage": 0, "lt_price": 0, "spot_price": 0}
    old = load_month_input(int(row.id), month)
    me_done = max(0.0, float(old.get("mechanism_energy") or 0.0))
    assess_actual_done = max(0.0, sm["actual"] - me_done)
    assess_cov_done = sm["lt"] / assess_actual_done if assess_actual_done else 0.0
    st.caption(f"截至已上传日清分：已发 {sm['actual']:.2f} MWh｜机制电量 {me_done:.2f} MWh｜考核基准 {assess_actual_done:.2f} MWh｜已签 {sm['lt']:.2f} MWh｜考核覆盖率 {assess_cov_done:.2%}")
    a, b, c = st.columns(3)
    p10 = a.number_input("剩余P10低发预测 MWh", 0.0); p50 = b.number_input("剩余P50预测 MWh", 0.0); p90 = c.number_input("剩余P90高发预测 MWh", 0.0)
    a, b, c, d = st.columns(4)
    remain_lt = a.number_input("剩余已签中长期 MWh", 0.0); expected_spot = b.number_input("预计剩余现货均价", value=float(sm['spot_price'])); trade_price = c.number_input("当前拟交易价格", value=float(sm['lt_price'])); proposed = d.number_input("试算交易量（卖出正/买回负）", value=0.0)
    risk_upper = float(row.get("risk_upper") or 1.10); total_lt = sm['lt'] + remain_lt + proposed; rows = []
    for name, rem in [("P10", p10), ("P50", p50), ("P90", p90)]:
        total_actual = sm['actual'] + rem; assess_total = max(0.0, total_actual - me_done); cov = total_lt / assess_total if assess_total else 0.0; space = assess_total * risk_upper - total_lt
        rows.append([name, total_actual, assess_total, total_lt, cov, space])
    x = pd.DataFrame(rows, columns=["情景", "预计月底总上网", "扣机制后考核电量", "交易后总合同", "考核仓位", "距风险上限空间"])
    st.dataframe(x, use_container_width=True, hide_index=True)
    spread = trade_price - expected_spot; p10space = float(x.iloc[0]["距风险上限空间"]); limit = float(row.get("trade_limit_mwh") or 100); threshold = float(row.get("min_spread") or 10)
    if p10space < 0: suggestion = f"优先买回/降仓，至少 {-p10space:.2f} MWh"
    elif spread >= threshold: suggestion = f"可考虑卖出，单次建议不超过 {min(p10space, limit):.2f} MWh"
    else: suggestion = "观望/保留现货"
    st.metric("中长期-预计现货价差", f"{spread:.2f} 元/MWh"); st.success(suggestion)

elif page == "数据管理":
    st.title("数据管理")
    row, _ = station_selector("data_station", False)
    if row is None: st.stop()
    df = daily_rows(int(row.id), 500); st.dataframe(df, use_container_width=True, hide_index=True)
    if not df.empty:
        day = st.selectbox("查看/删除日期", df.trade_date.astype(str).tolist()); rec = df[df.trade_date.astype(str) == day].iloc[0]
        with st.expander("24时点明细"): st.dataframe(hourly_rows(int(row.id), day), use_container_width=True, hide_index=True)
        if st.button("删除该日数据", type="secondary"):
            delete_daily(int(rec.id), int(row.id), day); st.success("已删除"); st.rerun()

elif page == "系统状态":
    st.title("系统状态")
    if DB_OK:
        st.success("Supabase 数据库已连接")
        st.write("规则版本：", RULE_VERSION)
        st.write("考核电量口径：Qactual - 机制电量 后再乘90%/110%")
        st.write("考核价格口径：P_LT=本站中长期均价；P_统一=P_节点=现货均价；P_区域同类型=本站中长期均价")
        st.write("绿电口径：Q绿电=min(Q绿电合约, Qactual-Q机制)；F绿电=Q绿电×绿色权益价格")
        st.write("自动结算展示：月度结算表 + 费用汇总表")
        st.write("核心表：stations / daily_summary / hourly_detail / monthly_settlement")
    else:
        st.error(DB_ERROR)
